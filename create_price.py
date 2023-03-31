import sys
import time
from multiprocessing import Process
from datetime import datetime
from time import sleep
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from PIL.PngImagePlugin import PngImageFile
from data_model import PriceItem, WarehouseItem, ShopItem, SmallWholesaleItem, MediumWholesaleItem, LargeWholesaleItem, warehouse_title_line, shop_title_line, small_wholesale_title_line, medium_wholesale_title_line, large_wholesale_title_line
from get_data_from_xlsx import get_data_from_xlsx
from fix_1c_error import fix_1c_error

def prepare_price(data: list[PriceItem], mode: str) -> list[tuple]:
    prepared_price = []
    match mode:
        case 'warehouse':
            prepared_price.append(warehouse_title_line)
            for item in data:   
                prepared_price.append(
                    WarehouseItem(
                    title=item.title,
                    warehouse_price=item.warehouse_price,
                    quantity_warehouse=item.quantity_warehouse,
                    quantity_shop1 = item.quantity_shop1,
                    quantity_shop2 = item.quantity_shop2,
                    quantity_shop3 = item.quantity_shop3,
                    img=item.img
                    )
                    )
        case 'shop':
            prepared_price.append(shop_title_line)
            for item in data:
                prepared_price.append(
                    ShopItem(
                    title=item.title,
                    shop_price=item.shop_price,
                    quantity_shop1 = item.quantity_shop1,
                    quantity_warehouse=item.quantity_warehouse,
                    quantity_shop2 = item.quantity_shop2,
                    quantity_shop3 = item.quantity_shop3,
                    img=item.img
                    )
                    )
        case 'small':
            prepared_price.append(small_wholesale_title_line)
            for item in data:
                if item.quantity_warehouse >= 0.1 and item.small_wholesale_price:
                    prepared_price.append(
                        SmallWholesaleItem(
                        title=item.title,
                        small_wholesale_price=item.small_wholesale_price,
                        img=item.img,
                        order = 0.0,
                        cost = None
                        )
                        )
        case 'medium':
            prepared_price.append(medium_wholesale_title_line)
            for item in data:
                if item.quantity_warehouse >= 0.1 and item.medium_wholesale_price:
                    prepared_price.append(
                        MediumWholesaleItem(
                        title=item.title,
                        medium_wholesale_price=item.medium_wholesale_price,
                        img=item.img,
                        order = 0.0,
                        cost = None
                        )
                        )
        case 'large':
            prepared_price.append(large_wholesale_title_line)
            for item in data:
                if item.quantity_warehouse >= 0.1 and item.large_wholesale_price:
                    prepared_price.append(
                        LargeWholesaleItem(
                        title=item.title,
                        large_wholesale_price=item.large_wholesale_price,
                        img=item.img,
                        order = 0.0,
                        cost = None
                        )
                        )
    return prepared_price


def paint_cell(ws: Worksheet, row: int, column: int, color: str) -> None:
    ws.cell(row=row, column=column).fill = PatternFill('solid', start_color=color)


def print_border(ws: Worksheet, row: int, column: int) -> None:
    ws.cell(row=row, column=column).border = Border(top=Side(border_style='thin', color='000000'),
                                                    left=Side(border_style='thin', color='000000'),
                                                    right=Side(border_style='thin', color='000000'),
                                                    bottom=Side(border_style='thin', color='000000')
                                                    )


def create_price(prepared_price: list[tuple]) -> Workbook:
    wb = Workbook()
    ws = wb.active

    warehouse_column = prepared_price[0].index('01.Склад') + 1
    shop1_column = prepared_price[0].index('02.ЦЧК') + 1
    shop2_column = prepared_price[0].index('03.Дом Чая') + 1
    shop3_column = prepared_price[0].index('04.Аллея') + 1


    for item_number, item in enumerate(prepared_price, 1):
        for field_number, field in enumerate(item, 1):

            print_border(ws, row=item_number, column=field_number)

            if field_number == warehouse_column:
                paint_cell(ws, row=item_number, column=field_number, color='C4D79B')
            elif field_number == shop1_column:
                paint_cell(ws, row=item_number, column=field_number, color='92CDDC')
            elif field_number == shop2_column:
                paint_cell(ws, row=item_number, column=field_number, color='FABF8F')
            elif field_number == shop3_column:
                paint_cell(ws, row=item_number, column=field_number, color='95B3D7')

            if isinstance(field, PngImageFile):
                ws.row_dimensions[item_number].height = 56 
                anchor = ws.cell(row=item_number, column=field_number).coordinate
                ws.add_image(Image(field), anchor=anchor)
            else:
                ws.cell(row=item_number, column=field_number).value = field

    ws.column_dimensions['A'].width = 90 #TODO: тут нужно сделать autofit для всех колонок
    return wb


def add_math(ws: Worksheet) -> None:
    last_row = ws.max_row
    for row_number, cell in enumerate(ws['E2': f'E{last_row}'], 2):
        cell[0].value = f'=PRODUCT(C{row_number} : D{row_number})'

    ws[f'E{last_row + 2}'].value = f'=SUM(E2 : E{last_row})'

def create_wholesale_price(prepared_price: list[tuple]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    for item_number, item in enumerate(prepared_price, 1):
        for field_number, field in enumerate(item, 1):
            print_border(ws, row=item_number, column=field_number)

            if isinstance(field, PngImageFile):
                ws.row_dimensions[item_number].height = 56 
                anchor = ws.cell(row=item_number, column=field_number).coordinate
                ws.add_image(Image(field), anchor=anchor)
            else:
                ws.cell(row=item_number, column=field_number).value = field
    ws.column_dimensions['A'].width = 90 #TODO: тут нужно сделать autofit для всех колонок
    add_math(ws)
    return wb


def main(path: str, mode: str, date: str) -> None:
    print(f'начинаем {mode}\n{"*"*80}')
    # date = datetime.now().strftime('%Y_%m_%d')

    # path = sys.argv[1]
    data = get_data_from_xlsx(path)
    prepared_price = prepare_price(data=data, mode=mode)
    name = f'{mode}_{date}.xlsx'
    if mode in ['warehouse', 'shop']:
        workbook = create_price(prepared_price)
    elif mode in ['small', 'medium', 'large']:
        workbook = create_wholesale_price(prepared_price)
    workbook.save(name)
    workbook.close()
    print(f'{mode} готов\n{"*"*80}')


if __name__ == '__main__':
    start_time = time.time()

    path = sys.argv[1]
    fix_1c_error(path)

    data = get_data_from_xlsx(path)
    date = datetime.now().strftime('%Y_%m_%d')
        
    processes = []

    for mode in ['warehouse', 'shop', 'small', 'medium', 'large']:
        processes.append(Process(target=main, args=(path, mode, date,), daemon=True))
        
    [p.start() for p in processes]
    [p.join() for p in processes]

    print(time.time() - start_time)